"""
Phase 1 ETL: extract the proposed Subdivision -> Studio -> Customer hierarchy
from the Personal-CRM operational workbook.

Hierarchy interpretation (confirmed):
  - our SUBDIVISION  = the publishing label/company (sheet section dividers, and
                       the "Studio" column in 'New Games POCs')
  - our STUDIO       = the sub-group within a label (a sheet's "Sub-studio:" label
                       and the "Subdivision" column in 'New Games POCs')
  - our CUSTOMER     = the individual game (one workbook sheet per game)

Output: data/hierarchy_proposed.csv  (review/edit this, then we load it).
This is best-effort; low-confidence rows are flagged in the `review` column.
"""

import csv
import os
import openpyxl

WB_PATH = "docs/Personal-CRM Data.xlsx"
OUT_PATH = "data/hierarchy_proposed.csv"

# Section-divider sheets ("X >>") that represent real publishing subdivisions.
# (Meta dividers Partners/Games/Prototypes/Admin are NOT customer subdivisions.)
SUBDIVISION_DIVIDERS = {
    "Vertex >>": "Vertex",
    "Ember Motion >>": "Ember Motion Games",
    "Nova >>": "Nova Studios",
    "Puzzlecraft >>": "Puzzlecraft",
    "Skyline Games >>": "Skyline Games",
    "Giantsworth >>": "Giantsworth Games",
    "Emerging Platforms >>": "Emerging Platforms",
    "Other T2 >>": "Other T2",
}

# Sheets that are NOT games (reference, internal partner orgs, studio-level logs).
NON_GAME_SHEETS = {
    "Cover", "Product Lookup", "Integration Tracker(pulled)", "New Games POCs",
    "Pod Info", "Pub Product", "CTO Strike", "PubOps", "Analytics", "VIP", "Mavens",
    # Studio-level note sheets (used as studios, not customers):
    "Atlas Studio", "NorthPeak Games",
}

# These sheets are studios (a sheet exists for the studio itself, with games nested).
STUDIO_LEVEL_SHEETS = {"Atlas Studio", "NorthPeak Games"}

APP_STATUS_HINTS = {
    "preproduction": "pre-production",
    "pre-production": "pre-production",
    "prototype": "prototype",
    "launched/live": "live-worldwide",
    "live": "live-worldwide",
    "soft launch": "soft-launch",
    "sunsetting": "sunsetting",
    "sunset": "sunset",
}


def norm(s):
    return (str(s).strip() if s is not None else "")


def first_rows(ws, n=6):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= n:
            break
    return rows


def load_new_games_pocs(wb):
    """game name -> (subdivision=Studio col, studio=Subdivision col)."""
    ws = wb["New Games POCs"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        game = norm(row[0])
        studio_col = norm(row[3])       # their "Studio"      -> our SUBDIVISION
        subdiv_col = norm(row[4])       # their "Subdivision" -> our STUDIO
        if game and studio_col:
            mapping.setdefault(game, (studio_col, subdiv_col))
    return mapping


def detect_status(rows):
    for row in rows:
        for cell in row[:3]:
            v = norm(cell).lower()
            if v in APP_STATUS_HINTS:
                return APP_STATUS_HINTS[v]
    return ""


def detect_substudio(rows):
    for row in rows:
        if norm(row[0]).lower().startswith("sub-studio"):
            return norm(row[1])
    return ""


def main():
    wb = openpyxl.load_workbook(WB_PATH, data_only=True, read_only=True)
    pocs = load_new_games_pocs(wb)

    os.makedirs("data", exist_ok=True)

    records = []
    current_sub = ""
    seen_games_divider = False

    for name in wb.sheetnames:
        if name == "Games >>":
            seen_games_divider = True
            continue
        if name.endswith(">>"):
            current_sub = SUBDIVISION_DIVIDERS.get(name, current_sub if name not in (
                "Partners >>", "Prototypes >>", "Admin >>") else current_sub)
            continue
        if not seen_games_divider:
            continue  # skip the Partners group at the top
        if name in NON_GAME_SHEETS:
            continue

        ws = wb[name]
        rows = first_rows(ws)

        subdivision = current_sub
        studio = detect_substudio(rows)
        status = detect_status(rows)
        review = []

        # Prefer the explicit New Games POCs mapping when present.
        if name in pocs:
            poc_sub, poc_studio = pocs[name]
            subdivision = poc_sub or subdivision
            studio = poc_studio or studio

        if not subdivision:
            subdivision = "UNKNOWN"
            review.append("no-subdivision")
        if not studio:
            studio = subdivision  # game sits directly under the label
            review.append("studio=subdivision(fallback)")
        if not status:
            status = "production"
            review.append("status-default")

        records.append({
            "customer": name,
            "subdivision": subdivision,
            "studio": studio,
            "app_status": status,
            "slack_channel": "",
            "review": ";".join(review),
        })

    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "customer", "subdivision", "studio", "app_status", "slack_channel", "review"])
        w.writeheader()
        w.writerows(records)

    # Console summary grouped by subdivision -> studio
    print(f"Wrote {len(records)} customers to {OUT_PATH}\n")
    tree = {}
    for r in records:
        tree.setdefault(r["subdivision"], {}).setdefault(r["studio"], []).append(r)
    for sub in tree:
        print(f"SUBDIVISION: {sub}")
        for studio in tree[sub]:
            print(f"   STUDIO: {studio}")
            for r in tree[sub][studio]:
                flag = f"   [{r['review']}]" if r["review"] else ""
                print(f"      - {r['customer']}  ({r['app_status']}){flag}")
        print()


if __name__ == "__main__":
    main()
