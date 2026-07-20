"""
Phase 2B extractor: parse per-game sheets into interactions + team notes.

Per decision:
  Meetings  -> interactions (type 'meeting')
  Feedback  -> interactions (type 'other')
  Notes     -> team_notes on the customer
  (POC(s) / Active Feature Adoptions / Misc notes: are ignored.)

Attendee / author free-text is matched against the people roster (data/people.csv
-> internal roster) and contacts (data/contacts.csv -> external). Unmatched fragments are
reported for review. Writes reviewable CSVs; does NOT touch the database.

Outputs:
  data/interactions.csv : customer,type,source,date,title,notes,internal_attendees,external_attendees,unmatched
  data/team_notes.csv   : customer,date,author,author_matched,context,notes
"""

import csv
import re
import sys
from datetime import datetime, date
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

WB_PATH = "docs/Personal-CRM Data.xlsx"

SECTION_BY_MARKER = {
    "meetings": "meetings",
    "notes": "notes",
    "feedback": "feedback",
    "poc(s)": "skip",
    "active feature adoptions": "skip",
    "misc notes:": "misc",
}
HEADER_FIRSTCOL = {"date", "added"}


def norm(s):
    return (str(s).strip() if s is not None else "")


def fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    s = norm(v)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return ""  # unparseable / blank


def load_roster():
    eng, con = {}, {}
    with open("data/people.csv", newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            n = r["name"].strip()
            if n:
                eng[n] = n
    with open("data/contacts.csv", newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            n = r["name"].strip()
            if n:
                con[n] = n
    return eng, con


def load_customers():
    names = []
    with open("data/hierarchy.csv", newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            c = r["customer"].strip()
            if c:
                names.append(c)
    return names


NOISE_WORDS = {
    "engs", "eng", "others", "other", "team", "teams", "group", "etc", "via",
    "slack", "tbd", "people", "dev", "devs", "folks", "and", "the", "plus",
}
EMAIL_RE = re.compile(r"[\w.\-+]+@[\w.\-]+")


def is_name_token(t):
    core = t.replace(".", "").replace("-", "").replace("'", "")
    return len(core) > 0 and all(c.isalpha() for c in core)


def _strip_known(text, names):
    """Remove any known names (case-insensitive) from text; return (found, remainder)."""
    found = []
    work = text
    for name in sorted(names, key=len, reverse=True):
        if re.search(re.escape(name), work, flags=re.IGNORECASE):
            found.append(name)
            work = re.sub(re.escape(name), " ", work, flags=re.IGNORECASE)
    return found, work


def parse_attendees(text, eng_names, con_names):
    """
    Delimiter-aware parse of an attendee/giver cell.
    Returns dict: internal, ext_existing, new[(name,email)], ambiguous, noise.
    """
    res = {"internal": [], "ext_existing": [], "new": [], "ambiguous": [], "noise": []}
    if not text:
        return res
    pieces = re.split(r"[\n,;]+", text)
    for piece in pieces:
        piece = piece.strip()
        if not piece:
            continue
        internal, rem = _strip_known(piece, eng_names)
        ext, rem = _strip_known(rem, con_names)
        res["internal"].extend(internal)
        res["ext_existing"].extend(ext)
        # what's left of the piece -> candidate new external name
        rem = re.sub(r"\(.*?\)", " ", rem)  # drop parentheticals e.g. (Skyline Games)
        emails = EMAIL_RE.findall(rem)
        rem = EMAIL_RE.sub(" ", rem)
        rem = re.sub(r"\s+", " ", rem).strip(" ,;:/&-.\t")
        if not rem:
            continue
        toks = rem.split()
        low = rem.lower()
        has_noise = any(t.lower() in NOISE_WORDS for t in toks) or re.search(r"\d", rem)
        looks_namey = all(is_name_token(t) for t in toks)
        if 2 <= len(toks) <= 3 and looks_namey and not has_noise:
            res["new"].append((rem, emails[0] if emails else ""))
        elif has_noise or len(toks) == 1:
            res["noise"].append(rem)
        else:
            res["ambiguous"].append(rem)
    return res


def main():
    eng_names, con_names = load_roster()
    customers = load_customers()
    cust_lookup = {c.lower(): c for c in customers}

    wb = openpyxl.load_workbook(WB_PATH, data_only=True, read_only=True)

    interactions = []  # rows for csv
    team_notes = []
    new_contacts = {}  # name(lower) -> (name, email, customer)
    per_customer = {}
    ambiguous_all = {}
    noise_all = {}
    misc_count = 0
    dateless = 0

    for sheet in wb.sheetnames:
        cust = cust_lookup.get(sheet.strip().lower())
        if not cust:
            continue  # not a customer sheet
        ws = wb[sheet]
        current = None
        counts = {"meetings": 0, "feedback": 0, "notes": 0}
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            a = norm(row[0]) if len(row) > 0 else ""
            al = a.lower()
            if al in SECTION_BY_MARKER:
                current = SECTION_BY_MARKER[al]
                continue
            if current in (None, "skip"):
                continue
            c0 = norm(row[0]) if len(row) > 0 else ""
            c1 = norm(row[1]) if len(row) > 1 else ""
            c2 = norm(row[2]) if len(row) > 2 else ""
            c3 = norm(row[3]) if len(row) > 3 else ""
            if current == "misc":
                if any([c0, c1, c2, c3]):
                    misc_count += 1
                continue
            # header row (Date/Added | Attendees/Added by/Feedback giver | Context | Notes) -> skip
            if (c2.lower() == "context"
                    or c0.lower() in HEADER_FIRSTCOL
                    or c1.lower() in ("attendees", "added by", "feedback giver")):
                continue

            d = fmt_date(row[0])
            people_text = c1
            title = c2
            body = c3

            # require real content; drop empty / junk rows
            if not (title or body):
                continue

            if current == "notes":
                pa = parse_attendees(people_text, eng_names, con_names)
                author = pa["internal"][0] if pa["internal"] else people_text
                team_notes.append([cust, d, author, "yes" if pa["internal"] else "no", title, body])
                counts["notes"] += 1
            else:
                if not d:
                    dateless += 1
                itype = "meeting" if current == "meetings" else "other"
                pa = parse_attendees(people_text, eng_names, con_names)
                for nm, em in pa["new"]:
                    key = nm.lower()
                    if key not in new_contacts:
                        new_contacts[key] = (nm, em, cust)
                    elif em and not new_contacts[key][1]:
                        prev = new_contacts[key]
                        new_contacts[key] = (prev[0], em, prev[2])
                for frag in pa["ambiguous"]:
                    ambiguous_all[frag] = ambiguous_all.get(frag, 0) + 1
                for frag in pa["noise"]:
                    noise_all[frag] = noise_all.get(frag, 0) + 1
                ext_all = pa["ext_existing"] + [nm for nm, _ in pa["new"]]
                unmatched = "; ".join(pa["ambiguous"] + pa["noise"])
                interactions.append([
                    cust, itype, current, d, title, body,
                    "; ".join(pa["internal"]), "; ".join(ext_all), unmatched,
                ])
                counts[current] += 1
        if any(counts.values()):
            per_customer[cust] = counts

    with open("data/interactions.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["customer", "type", "source", "date", "title", "notes",
                    "internal_attendees", "external_attendees", "unmatched"])
        w.writerows(interactions)
    with open("data/team_notes.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["customer", "date", "author", "author_matched", "context", "notes"])
        w.writerows(team_notes)
    with open("data/new_contacts.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "email", "customer"])
        for _, (nm, em, cust) in sorted(new_contacts.items()):
            w.writerow([nm, em, cust])

    print(f"Interactions: {len(interactions)}  |  Team notes: {len(team_notes)}  "
          f"|  New external contacts: {len(new_contacts)}  |  rows w/o date: {dateless}  |  Misc skipped: {misc_count}\n")
    print(f"{'CUSTOMER':24} {'meetings':>9} {'feedback':>9} {'notes':>7}")
    for cust, c in sorted(per_customer.items()):
        print(f"{cust:24} {c['meetings']:>9} {c['feedback']:>9} {c['notes']:>7}")
    print("\nNEW external contacts to be created (name -> customer [email]):")
    for _, (nm, em, cust) in sorted(new_contacts.items()):
        print(f"  + {nm}  ->  {cust}" + (f"  [{em}]" if em else ""))
    if ambiguous_all:
        print("\nAMBIGUOUS fragments (NOT auto-created -- multiple names jammed together, review later):")
        for frag, n in sorted(ambiguous_all.items(), key=lambda kv: -kv[1]):
            print(f"  ? [{n}x] {frag}")
    if noise_all:
        print("\nNoise fragments (ignored):")
        for frag, n in sorted(noise_all.items(), key=lambda kv: -kv[1]):
            print(f"  - [{n}x] {frag}")


if __name__ == "__main__":
    main()
